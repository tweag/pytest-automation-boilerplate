"""
----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------

----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------
"""

import os
import csv
import structlog
import openpyxl

from openpyxl import utils
from pathlib import Path
from openpyxl.reader.excel import load_workbook
from pytest_bdd import parsers, given, when, then
from bp_core.frontend.common.helpers.app import context_manager
from bp_core.frontend.common.helpers.selenium_generics import SeleniumGenerics
from bp_core.frontend.common.step_definitions.steps_common import MOBILE_SUFFIX
from bp_core.frontend.common.utils.locator_parser import Locators
from bp_core.utils import data_manager
from bp_core.utils.bp_storage import BPStorage

logger = structlog.get_logger(__name__)


# WEB & MOBILE contexts Predefined Step
# ID 801
@given(parsers.re("I get text from element '(?P<locator_path>.*)' and save as environment variable '(?P<env_var>.*)'"))
@when(parsers.re("I get text from element '(?P<locator_path>.*)' and save as environment variable '(?P<env_var>.*)'"))
def get_text_from_element(selenium_generics: SeleniumGenerics, locators: Locators, locator_path: str, env_var: str, ):
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            returned_text = selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))
    else:
        returned_text = selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))
    os.environ[env_var] = returned_text


# ID 802
@given(parsers.re("I get text from element '(?P<locator_path>.*)' between '(?P<initial_string>.*)' and '(?P<final_string>.*)' boundaries, and save as environment variable '(?P<env_var>.*)'"),
    converters=dict(initial_string=data_manager.text_formatted, final_string=data_manager.text_formatted), )
@when(parsers.re("I get text from element '(?P<locator_path>.*)' between '(?P<initial_string>.*)' and '(?P<final_string>.*)' boundaries, and save as environment variable '(?P<env_var>.*)'"),
    converters=dict(initial_string=data_manager.text_formatted, final_string=data_manager.text_formatted), )
def get_text_from_element(selenium_generics: SeleniumGenerics, locators: Locators, locator_path: str, initial_string: str, final_string:str, env_var: str, ):
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            returned_text = selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))
    else:
        returned_text = selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))
    os.environ[env_var] = returned_text[returned_text.index(initial_string) + len(initial_string): returned_text.index(final_string)]


# WEB context Predefined Step
# ID 803
@given(parsers.re("I write within the HTML report the environment variable '(?P<env_var>.*)' value"))
@when(parsers.re("I write within the HTML report the environment variable '(?P<env_var>.*)' value"))
@then(parsers.re("I write within the HTML report the environment variable '(?P<env_var>.*)' value"))
def write_html_report_os_environ_value(selenium_generics: SeleniumGenerics, env_var: str):
    if BPStorage.get_env_vars_for_html() is not None:
        BPStorage.save_env_vars_for_html({**BPStorage.get_env_vars_for_html(), **{env_var: os.environ.get(env_var)}})
    else:
        BPStorage.save_env_vars_for_html({**{env_var: os.environ.get(env_var)}})


# WEB context Predefined Step
# ID 804
@given(parsers.re("I store '(?P<key>.*)' environment variable in .local.env config file"))
@when(parsers.re("I store '(?P<key>.*)' environment variable in .local.env config file"))
def store_env_variable_in_local_env(key: str):
    local_config_file = Path.cwd() / "configs" / ".local.env"
    if local_config_file.is_file():
        from dotenv import set_key
        if os.environ.get(key, None):
            value = os.environ[key]
            set_key(local_config_file.as_posix(), key, value)
        else:
            raise KeyError(f"Environment variable: {key} does not exits")
    else:
        raise FileNotFoundError("File not found: .local.env")


# ID 806, 909
@given(parsers.re("I get text from '(?P<cell>.*)' on '(?P<sheet_name>.*)' of excel file '(?P<file_path>.*)' and save it as environment variable with name '(?P<env_var>.*)'"),
       converters=dict(cell=data_manager.text_formatted, sheet_name=data_manager.text_formatted))
@when(parsers.re("I get text from '(?P<cell>.*)' on '(?P<sheet_name>.*)' of excel file '(?P<file_path>.*)' and save it as environment variable with name '(?P<env_var>.*)'"),
      converters=dict(cell=data_manager.text_formatted, sheet_name=data_manager.text_formatted))
def write_text_to_excel_file(cell: str, sheet_name: str, file_path: str, env_var: str):
    sheet = load_workbook(file_path)[sheet_name]
    text = sheet[cell].value
    os.environ[env_var] = text


# ID 807, 910
@given(parsers.re("I get text of '(?P<sheet_name>.*)' of excel file '(?P<file_path>.*)' and save it as environment variables"),
       converters=dict(sheet_name=data_manager.text_formatted))
@when(parsers.re("I get text of '(?P<sheet_name>.*)' of excel file '(?P<file_path>.*)' and save it as environment variables"),
      converters=dict(sheet_name=data_manager.text_formatted))
def store_excel_data_as_env_vars(sheet_name: str, file_path: str):
    sheet = load_workbook(file_path)[sheet_name]
    first_column = None
    min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(sheet.dimensions)
    for col in sheet.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row, values_only=False):
        if first_column:
            break
        for cell in col:
            if cell.value:
                first_column = cell.column
                break
    else:
        raise ValueError(f"No data found in {file_path}")
    for key, value in sheet.iter_rows(min_col=first_column, max_col=first_column + 1, min_row=min_row, max_row=max_row, values_only=True):
        if key and value:
            os.environ[key] = str(value)


# ID 808, 920
@given(parsers.re("I get text from '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' and save it as env variable with name '(?P<env_var>.*)'"),
       converters=dict(cell=data_manager.text_formatted))
@when(parsers.re("I get text from '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' and save it as env variable with name '(?P<env_var>.*)'"),
       converters=dict(cell=data_manager.text_formatted))
def save_csv_cell_text_as_environment_variable(cell: str, file_path: str, env_var: str):
    csv_file = Path(file_path).absolute()
    if not all([csv_file.exists(), csv_file.is_file(), csv_file.suffix == '.csv']):
        raise FileNotFoundError(f"File {file_path} is not a valid csv file")
    wb = openpyxl.Workbook()
    sheet = wb.active
    with open(csv_file, newline="") as f:
        _csv_file = f.read()
    dialect = csv.Sniffer()
    delimiter = dialect.sniff(_csv_file).delimiter
    with open(csv_file, newline="") as file:
        reader = csv.reader(file, delimiter=delimiter)
        for row in reader:
            sheet.append(row)

    text = sheet[cell].value
    os.environ[env_var] = text
