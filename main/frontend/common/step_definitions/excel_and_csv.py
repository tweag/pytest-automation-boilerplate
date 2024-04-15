import os
import csv
import structlog
import openpyxl

from pathlib import Path
from openpyxl.reader.excel import load_workbook
from pytest_bdd import parsers, given, when, then
from assertpy import assert_that
from main.utils import data_manager

logger = structlog.get_logger(__name__)


@given(parsers.re("Text inside '(?P<cell>.+)' on sheet '(?P<sheet_name>.+)' of excel file '(?P<file_path>.+)' is equal to '(?P<expected_text>.+)'"),
      converters=dict(cell=data_manager.text_formatted, sheet_name=data_manager.text_formatted, expected_text=data_manager.text_formatted))
@when(parsers.re("Text inside '(?P<cell>.+)' on sheet '(?P<sheet_name>.+)' of excel file '(?P<file_path>.+)' is equal to '(?P<expected_text>.+)'"),
      converters=dict(cell=data_manager.text_formatted, sheet_name=data_manager.text_formatted, expected_text=data_manager.text_formatted))
@then(parsers.re("Text inside '(?P<cell>.+)' on sheet '(?P<sheet_name>.+)' of excel file '(?P<file_path>.+)' is equal to '(?P<expected_text>.+)'"),
      converters=dict(cell=data_manager.text_formatted, sheet_name=data_manager.text_formatted, expected_text=data_manager.text_formatted))
def cell_text_is_equal_to(cell: str, sheet_name: str, file_path: str, expected_text: str):
    sheet = load_workbook(file_path)[sheet_name]
    assert_that(sheet[cell].value).is_equal_to(expected_text)


# ID 902
@given(parsers.re("Text inside '(?P<cell>.+)' on sheet '(?P<sheet_name>.+)' of excel file '(?P<file_path>.+)' contains '(?P<expected_text>.+)'"),
      converters=dict(cell=data_manager.text_formatted, sheet_name=data_manager.text_formatted, expected_text=data_manager.text_formatted))
@when(parsers.re("Text inside '(?P<cell>.+)' on sheet '(?P<sheet_name>.+)' of excel file '(?P<file_path>.+)' contains '(?P<expected_text>.+)'"),
      converters=dict(cell=data_manager.text_formatted, sheet_name=data_manager.text_formatted, expected_text=data_manager.text_formatted))
@then(parsers.re("Text inside '(?P<cell>.+)' on sheet '(?P<sheet_name>.+)' of excel file '(?P<file_path>.+)' contains '(?P<expected_text>.+)'"),
      converters=dict(cell=data_manager.text_formatted, sheet_name=data_manager.text_formatted, expected_text=data_manager.text_formatted))
def cell_text_contains(cell:str, sheet_name: str, file_path: str, expected_text: str):
    sheet = load_workbook(file_path)[sheet_name]
    assert_that(sheet[cell].value).contains(expected_text)


@given(parsers.re("The '(?P<cell>.+)' on sheet '(?P<sheet_name>.+)' of excel file '(?P<file_path>.+)' does not contain any text"),
      converters=dict(cell=data_manager.text_formatted, sheet_name=data_manager.text_formatted))
@when(parsers.re("The '(?P<cell>.+)' on sheet '(?P<sheet_name>.+)' of excel file '(?P<file_path>.+)' does not contain any text"),
      converters=dict(cell=data_manager.text_formatted, sheet_name=data_manager.text_formatted))
@then(parsers.re("The '(?P<cell>.+)' on sheet '(?P<sheet_name>.+)' of excel file '(?P<file_path>.+)' does not contain any text"),
      converters=dict(cell=data_manager.text_formatted, sheet_name=data_manager.text_formatted))
def cell_text_is_empty(cell:str, sheet_name: str, file_path: str):
    sheet = load_workbook(file_path)[sheet_name]
    assert_that(sheet[cell].value).is_none()


@given(parsers.re("The '(?P<cell>.+)' on sheet '(?P<sheet_name>.+)' of excel file '(?P<file_path>.+)' text is equal with the text of the '(?P<locator_path>.+)'"),
      converters=dict(cell=data_manager.text_formatted, sheet_name=data_manager.text_formatted))
@when(parsers.re("The '(?P<cell>.+)' on sheet '(?P<sheet_name>.+)' of excel file '(?P<file_path>.+)' text is equal with the text of the '(?P<locator_path>.+)'"),
      converters=dict(cell=data_manager.text_formatted, sheet_name=data_manager.text_formatted))
@then(parsers.re("The '(?P<cell>.+)' on sheet '(?P<sheet_name>.+)' of excel file '(?P<file_path>.+)' text is equal with the text of the '(?P<locator_path>.+)'"),
      converters=dict(cell=data_manager.text_formatted, sheet_name=data_manager.text_formatted))
def cell_text_equals_element_text(selenium_generics, locators, cell:str, sheet_name: str, file_path: str, locator_path: str):
    sheet = load_workbook(file_path)[sheet_name]
    assert_that(sheet[cell].value).is_equal_to(selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics)))


@given(parsers.re("The '(?P<cell>.+)' on sheet '(?P<sheet_name>.+)' of excel file '(?P<file_path>.+)' is contained in the text of '(?P<locator_path>.+)'"),
      converters=dict(cell=data_manager.text_formatted, sheet_name=data_manager.text_formatted))
@when(parsers.re("The '(?P<cell>.+)' on sheet '(?P<sheet_name>.+)' of excel file '(?P<file_path>.+)' is contained in the text of '(?P<locator_path>.+)'"),
      converters=dict(cell=data_manager.text_formatted, sheet_name=data_manager.text_formatted))
@then(parsers.re("The '(?P<cell>.+)' on sheet '(?P<sheet_name>.+)' of excel file '(?P<file_path>.+)' is contained in the text of '(?P<locator_path>.+)'"),
      converters=dict(cell=data_manager.text_formatted, sheet_name=data_manager.text_formatted))
def element_text_contains_cell_text(selenium_generics, locators, cell:str, sheet_name: str, file_path: str, locator_path: str):
    sheet = load_workbook(file_path)[sheet_name]
    assert_that(selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))).contains(sheet[cell].value)


@given(parsers.re("I delete '(?P<file_path>.*)' file"))
@when(parsers.re("I delete '(?P<file_path>.*)' file"))
def delete_file(file_path: str):
    absolute_path = Path(file_path).absolute()
    if absolute_path.exists():
        os.remove(absolute_path)


# ID 907
@given(parsers.re("I create excel file '(?P<file_name>.*)' and save on '(?P<file_path>.+)'"))
@when(parsers.re("I create excel file '(?P<file_name>.*)' and save on '(?P<file_path>.+)'"))
def create_excel_file(file_name: str, file_path: str):
    if Path(file_name).suffix != '.xlsx':
        raise ValueError(f"Invalid file extension for {file_name}")
    excel_file = (Path(file_path) / file_name).absolute()
    openpyxl.Workbook().save(excel_file.as_posix())


@given(parsers.re("I write '(?P<text>.*)' to '(?P<cell>.*)' on '(?P<sheet_name>.*)' of excel file '(?P<file_path>.*)'"),
       converters=dict(cell=data_manager.text_formatted, sheet_name=data_manager.text_formatted,
       text=data_manager.text_formatted))
@when(parsers.re("I write '(?P<text>.*)' to '(?P<cell>.*)' on '(?P<sheet_name>.*)' of excel file '(?P<file_path>.*)'"),
      converters=dict(cell=data_manager.text_formatted, sheet_name=data_manager.text_formatted,
      text=data_manager.text_formatted))
def write_text_to_excel_file(text: str, cell: str, sheet_name: str, file_path: str):
    excel_file = Path(file_path).absolute().as_posix()
    wb = load_workbook(excel_file)
    wb[sheet_name][cell] = text
    wb.save(excel_file)


@given(parsers.re("Number of total rows on '(?P<sheet_name>.*)' of excel file '(?P<file_path>.*)' is '(?P<row_count>.*)'"),
      converters=dict(row_count=data_manager.text_formatted, sheet_name=data_manager.text_formatted))
@when(parsers.re("Number of total rows on '(?P<sheet_name>.*)' of excel file '(?P<file_path>.*)' is '(?P<row_count>.*)'"),
      converters=dict(row_count=data_manager.text_formatted, sheet_name=data_manager.text_formatted))
@then(parsers.re("Number of total rows on '(?P<sheet_name>.*)' of excel file '(?P<file_path>.*)' is '(?P<row_count>.*)'"),
      converters=dict(row_count=data_manager.text_formatted, sheet_name=data_manager.text_formatted))
def total_rows_number_with_data_is_equal_to(sheet_name: str, file_path: str, row_count: str):
    sheet = load_workbook(file_path)[sheet_name]
    num_rows = 0
    for row in sheet.iter_rows():
        if any(cell.value is not None for cell in row):
            num_rows += 1
    assert_that(int(num_rows)).is_equal_to(int(row_count))


@given(parsers.re("Number of rows containing '(?P<expected_text>.*)' on '(?P<sheet_name>.*)' of excel file '(?P<file_path>.*)' is '(?P<row_count>.*)'"),
      converters=dict(row_count=data_manager.text_formatted, sheet_name=data_manager.text_formatted, text=data_manager.text_formatted))
@when(parsers.re("Number of rows containing '(?P<expected_text>.*)' on '(?P<sheet_name>.*)' of excel file '(?P<file_path>.*)' is '(?P<row_count>.*)'"),
      converters=dict(row_count=data_manager.text_formatted, sheet_name=data_manager.text_formatted, text=data_manager.text_formatted))
@then(parsers.re("Number of rows containing '(?P<expected_text>.*)' on '(?P<sheet_name>.*)' of excel file '(?P<file_path>.*)' is '(?P<row_count>.*)'"),
      converters=dict(row_count=data_manager.text_formatted, sheet_name=data_manager.text_formatted, text=data_manager.text_formatted))
def number_rows_with_text_is_equal_to(expected_text: str, sheet_name: str, file_path: str, row_count: str):
    sheet = load_workbook(file_path)[sheet_name]
    num_rows = 0
    for row in sheet.iter_rows():
        if any(cell.value == f"{expected_text}" for cell in row):
            num_rows += 1
    assert_that(int(num_rows)).is_equal_to(int(row_count))


@given(parsers.re("Text inside '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' is equal to '(?P<expected_text>.+)'"),
      converters=dict(expected_text=data_manager.text_formatted, cell=data_manager.text_formatted))
@when(parsers.re("Text inside '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' is equal to '(?P<expected_text>.+)'"),
      converters=dict(expected_text=data_manager.text_formatted, cell=data_manager.text_formatted))
@then(parsers.re("Text inside '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' is equal to '(?P<expected_text>.+)'"),
      converters=dict(expected_text=data_manager.text_formatted, cell=data_manager.text_formatted))
def csv_cell_text_is_equal_to(cell: str, file_path: str, expected_text: str):
    csv_file = Path(file_path).absolute()
    if not all([csv_file.exists(), csv_file.is_file(), csv_file.suffix == '.csv']):
        raise FileNotFoundError(f"File {file_path} is not a valid csv file")
    wb = openpyxl.Workbook()
    sheet = wb.active
    with open(csv_file, newline="") as f:
        _csv_file = f.read()
    dialect = csv.Sniffer()
    delimiter = dialect.sniff(_csv_file).delimiter
    with open(csv_file, newline="") as file:
        reader = csv.reader(file, delimiter=delimiter)
        for row in reader:
            sheet.append(row)
    assert_that(sheet[cell].value).is_equal_to(expected_text)


@given(parsers.re("Text inside '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' contains '(?P<expected_text>.+)'"),
      converters=dict(expected_text=data_manager.text_formatted, cell=data_manager.text_formatted))
@when(parsers.re("Text inside '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' contains '(?P<expected_text>.+)'"),
      converters=dict(expected_text=data_manager.text_formatted, cell=data_manager.text_formatted))
@then(parsers.re("Text inside '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' contains '(?P<expected_text>.+)'"),
      converters=dict(expected_text=data_manager.text_formatted, cell=data_manager.text_formatted))
def csv_cell_text_contains(cell: str, file_path: str, expected_text: str):
    csv_file = Path(file_path).absolute()
    if not all([csv_file.exists(), csv_file.is_file(), csv_file.suffix == '.csv']):
        raise FileNotFoundError(f"File {file_path} is not a valid csv file")
    wb = openpyxl.Workbook()
    sheet = wb.active
    with open(csv_file, newline="") as f:
        _csv_file = f.read()
    dialect = csv.Sniffer()
    delimiter = dialect.sniff(_csv_file).delimiter
    with open(csv_file, newline="") as file:
        reader = csv.reader(file, delimiter=delimiter)
        for row in reader:
            sheet.append(row)
    assert_that(sheet[cell].value).contains(expected_text)


@given(parsers.re("The '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' text is equal with the text of the '(?P<locator_path>.+)'"),
      converters=dict(cell=data_manager.text_formatted))
@when(parsers.re("The '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' text is equal with the text of the '(?P<locator_path>.+)'"),
      converters=dict(cell=data_manager.text_formatted))
@then(parsers.re("The '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' text is equal with the text of the '(?P<locator_path>.+)'"),
      converters=dict(cell=data_manager.text_formatted))
def csv_cell_text_equals_element_text(selenium_generics, locators, cell: str, file_path: str, locator_path: str):
    csv_file = Path(file_path).absolute()
    if not all([csv_file.exists(), csv_file.is_file(), csv_file.suffix == '.csv']):
        raise FileNotFoundError(f"File {file_path} is not a valid csv file")
    wb = openpyxl.Workbook()
    sheet = wb.active
    with open(csv_file, newline="") as f:
        _csv_file = f.read()
    dialect = csv.Sniffer()
    delimiter = dialect.sniff(_csv_file).delimiter
    with open(csv_file, newline="") as file:
        reader = csv.reader(file, delimiter=delimiter)
        for row in reader:
            sheet.append(row)
    assert_that(sheet[cell].value).is_equal_to(selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics)))


@given(parsers.re("The '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' is contained in the text of '(?P<locator_path>.+)'"),
      converters=dict(cell=data_manager.text_formatted))
@when(parsers.re("The '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' is contained in the text of '(?P<locator_path>.+)'"),
      converters=dict(cell=data_manager.text_formatted))
@then(parsers.re("The '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' is contained in the text of '(?P<locator_path>.+)'"),
      converters=dict(cell=data_manager.text_formatted))
def element_text_contains_csv_cell_text(selenium_generics, locators, cell: str, file_path: str, locator_path: str):
    csv_file = Path(file_path).absolute()
    if not all([csv_file.exists(), csv_file.is_file(), csv_file.suffix == '.csv']):
        raise FileNotFoundError(f"File {file_path} is not a valid csv file")
    wb = openpyxl.Workbook()
    sheet = wb.active
    with open(csv_file, newline="") as f:
        _csv_file = f.read()
    dialect = csv.Sniffer()
    delimiter = dialect.sniff(_csv_file).delimiter
    with open(csv_file, newline="") as file:
        reader = csv.reader(file, delimiter=delimiter)
        for row in reader:
            sheet.append(row)
    assert_that(selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))).contains(
        sheet[cell].value)


@given(parsers.re("The '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' does not contain any text"),
      converters=dict(cell=data_manager.text_formatted))
@when(parsers.re("The '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' does not contain any text"),
      converters=dict(cell=data_manager.text_formatted))
@then(parsers.re("The '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' does not contain any text"),
      converters=dict(cell=data_manager.text_formatted))
def csv_cell_text_is_empty(cell: str, file_path: str):
    csv_file = Path(file_path).absolute()
    if not all([csv_file.exists(), csv_file.is_file(), csv_file.suffix == '.csv']):
        raise FileNotFoundError(f"File {file_path} is not a valid csv file")
    wb = openpyxl.Workbook()
    sheet = wb.active
    with open(csv_file, newline="") as f:
        _csv_file = f.read()
    dialect = csv.Sniffer()
    delimiter = dialect.sniff(_csv_file).delimiter
    with open(csv_file, newline="") as file:
        reader = csv.reader(file, delimiter=delimiter)
        for row in reader:
            sheet.append(row)
    assert_that(sheet[cell].value).is_empty()


@given(parsers.re("Number of rows containing '(?P<expected_text>.*)' of csv file '(?P<file_path>.*)' is '(?P<row_count>.*)'"),
      converters=dict(expected_text=data_manager.text_formatted, row_count=data_manager.text_formatted))
@when(parsers.re("Number of rows containing '(?P<expected_text>.*)' of csv file '(?P<file_path>.*)' is '(?P<row_count>.*)'"),
      converters=dict(expected_text=data_manager.text_formatted, row_count=data_manager.text_formatted))
@then(parsers.re("Number of rows containing '(?P<expected_text>.*)' of csv file '(?P<file_path>.*)' is '(?P<row_count>.*)'"),
      converters=dict(expected_text=data_manager.text_formatted, row_count=data_manager.text_formatted))
def number_csv_rows_with_text_is_equal_to(expected_text: str, file_path: str, row_count: str):
    csv_file = Path(file_path).absolute()
    if not all([csv_file.exists(), csv_file.is_file(), csv_file.suffix == '.csv']):
        raise FileNotFoundError(f"File {file_path} is not a valid csv file")
    wb = openpyxl.Workbook()
    sheet = wb.active
    with open(csv_file, newline="") as f:
        _csv_file = f.read()
    dialect = csv.Sniffer()
    delimiter = dialect.sniff(_csv_file).delimiter
    with open(csv_file, newline="") as file:
        reader = csv.reader(file, delimiter=delimiter)
        for row in reader:
            sheet.append(row)

    num_rows = 0
    for row in sheet.iter_rows():
        if any(cell.value == f"{expected_text}" for cell in row):
            num_rows += 1
    assert_that(int(num_rows)).is_equal_to(int(row_count))


@given(parsers.re("Number of total rows of csv file '(?P<file_path>.*)' is '(?P<row_count>.*)'"),
      converters=dict(row_count=data_manager.text_formatted))
@when(parsers.re("Number of total rows of csv file '(?P<file_path>.*)' is '(?P<row_count>.*)'"),
      converters=dict(row_count=data_manager.text_formatted))
@then(parsers.re("Number of total rows of csv file '(?P<file_path>.*)' is '(?P<row_count>.*)'"),
      converters=dict(row_count=data_manager.text_formatted))
def total_csv_rows_number_with_data_is_equal_to(file_path: str, row_count: str):
    csv_file = Path(file_path).absolute()
    if not all([csv_file.exists(), csv_file.is_file(), csv_file.suffix == '.csv']):
        raise FileNotFoundError(f"File {file_path} is not a valid csv file")
    wb = openpyxl.Workbook()
    sheet = wb.active
    with open(csv_file, newline="") as f:
        _csv_file = f.read()
    dialect = csv.Sniffer()
    delimiter = dialect.sniff(_csv_file).delimiter
    with open(csv_file, newline="") as file:
        reader = csv.reader(file, delimiter=delimiter)
        for row in reader:
            sheet.append(row)

    num_rows = 0
    for row in sheet.iter_rows():
        if any(cell.value is not None for cell in row):
            num_rows += 1
    assert_that(int(num_rows)).is_equal_to(int(row_count))


@given(parsers.re("I Write '(?P<text>.+)' to '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)'"),
       converters=dict(cell=data_manager.text_formatted, text=data_manager.text_formatted))
@when(parsers.re("I Write '(?P<text>.+)' to '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)'"),
       converters=dict(cell=data_manager.text_formatted, text=data_manager.text_formatted))
def write_text_to_csv_cell(text: str, cell: str, file_path: str):
    csv_file = Path(file_path).absolute()
    wb = openpyxl.Workbook()
    sheet = wb.active
    delimiter = None

    if csv_file.exists():
        with open(csv_file, newline="") as f:
            _csv_file = f.read()
        if _csv_file:
            dialect = csv.Sniffer()
            delimiter = dialect.sniff(_csv_file).delimiter
    else:
        with open(csv_file, "w", newline="") as f:
            f.write("")  # Creating the file as it does not exist
    delimiter = delimiter if delimiter else ','

    with open(csv_file, newline="") as f:
        reader = csv.reader(f, delimiter=delimiter)
        for row in reader:
            sheet.append(row)
        sheet[cell] = text

    with open(csv_file, 'w', newline="") as f:
        writer = csv.writer(f, delimiter=delimiter)
        for row in sheet.iter_rows():
            writer.writerow([cell.value for cell in row])


@given(parsers.re("I create csv file '(?P<file_name>.*)' and save on '(?P<file_path>.+)'"))
@when(parsers.re("I create csv file '(?P<file_name>.*)' and save on '(?P<file_path>.+)'"))
def create_csv_file(file_name: str, file_path: str):
    if Path(file_name).suffix != '.csv':
        raise ValueError(f"Invalid file extension for {file_name}")
    csv_file = (Path(file_path) / file_name).absolute()

    # Create an empty csv file
    with open(csv_file, "w", newline="") as f:
        f.write("")
